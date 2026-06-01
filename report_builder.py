from __future__ import annotations

import html
import json
from pathlib import Path

import pandas as pd


class ReportBuilder:
    """Generate standalone HTML reports and full-detail Excel workbooks."""

    def __init__(self, analysis_result: dict):
        self.data = self._enrich_result(dict(analysis_result))

    def _enrich_result(self, data: dict) -> dict:
        materials = data.get("material_forecasts") or []
        procurement = data.get("procurement_suggestions") or []
        material_forecast_total = round(sum(float(x.get("forecast_total", 0) or 0) for x in materials), 1)
        data.setdefault("summary", {})["material_forecast_total"] = material_forecast_total
        data["summary"]["forecast_total"] = material_forecast_total

        if not data.get("forecast_months"):
            data["forecast_months"] = self._infer_forecast_months(data)
        data["rebuilt_from_legacy_json"] = not data.get("method_notes") or not data.get("forecast_months")
        data["total_forecast"] = self._normalize_total_forecast(data)

        if not data.get("trend_summary"):
            rows = []
            for label in ("上升", "平稳", "下降"):
                subset = [m for m in materials if label in str(m.get("trend", ""))]
                rows.append(
                    {
                        "trend": label,
                        "material_count": len(subset),
                        "total_12m": round(sum(float(m.get("total_12m", 0) or 0) for m in subset), 2),
                        "forecast_total": round(sum(float(m.get("forecast_total", 0) or 0) for m in subset), 2),
                    }
                )
            data["trend_summary"] = rows

        if not data.get("procurement_summary_by_priority"):
            data["procurement_summary_by_priority"] = self._group_procurement(procurement, "priority")
        if not data.get("procurement_summary_by_action"):
            data["procurement_summary_by_action"] = self._group_procurement(procurement, "purchase_timing")
        if not data.get("method_notes"):
            data["method_notes"] = {
                "forecast": [
                    "单物料预测采用保守集成预测：Holt趋势25% + 近3月均值55% + 长期月均20%。",
                    "Holt趋势使用双指数平滑：先维护水平值 level 和趋势值 trend。每个月更新 level = α×本月消耗 + (1-α)×(上月level+上月trend)，trend = β×(level-上月level) + (1-β)×上月trend。本报告取 α=0.25、β=0.10，再用 level + trend×未来期数 得到未来月份趋势值。",
                    "近3月均值反映最新业务节奏，长期月均用于防止短期波动过度影响预测，Holt用于识别趋势。",
                    "如果近6个月仍有消耗，则未来单月预测不低于近3月均值的50%，避免仍在消耗的物料被趋势外推压到0。",
                    "如果近6个月完全无消耗，才允许预测降至0。",
                ],
                "procurement": [
                    "日均消耗 = 月均消耗 / 30。",
                    "可撑天数 = 当前库存 / 日均消耗。",
                    "安全库存 = max(|近3月均值 - 月均消耗|, 月均消耗25%) × 1.64，约对应90%服务水平。",
                    "再订点 = 月均消耗 × 1个月采购提前期 + 安全库存。",
                    "建议采购量 = max(0, 未来3个月预测消耗 + 安全库存 - 当前库存)，并按10的倍数向上取整。",
                    "采购标签优先由建议采购量决定：建议采购量为0时不显示立即采购；若库存风险很高但预测需求不足，则标记为人工复核。",
                ],
            }
        if not data.get("inactive_material_summary"):
            data["inactive_material_summary"] = {"material_count": 0, "stocked_count": 0, "end_qty": 0, "end_amt": 0}
        if not data.get("inactive_materials"):
            data["inactive_materials"] = []
        for item in materials:
            if item.get("recent_6m_total") in (None, ""):
                try:
                    item["recent_6m_total"] = round(float(item.get("avg_monthly", 0) or 0) * 6, 1)
                except Exception:
                    item["recent_6m_total"] = 0
        return data

    def _normalize_total_forecast(self, data: dict) -> dict:
        total = dict(data.get("total_forecast") or {})
        months = data.get("forecast_months") or self._infer_forecast_months(data)
        if not months:
            return total
        if "recent_ma" not in total:
            total["recent_ma"] = total.get("ma_trend") or self._flat_forecast(data.get("summary", {}).get("forecast_avg_monthly", 0), months)
        if "long_avg" not in total:
            avg = data.get("summary", {}).get("avg_monthly", 0)
            total["long_avg"] = self._flat_forecast(avg, months)
        return total

    def _flat_forecast(self, value, months: list[str]) -> dict:
        try:
            value = float(value or 0)
        except Exception:
            value = 0
        return {month: round(value, 1) for month in months}

    def _infer_forecast_months(self, data: dict) -> list[str]:
        for key in ("ensemble", "holt", "recent_ma", "long_avg", "seasonal_naive", "ma_trend"):
            months = list((data.get("total_forecast") or {}).get(key, {}).keys())
            if months:
                return sorted(months)
        for item in data.get("material_forecasts") or []:
            months = list((item.get("forecast") or {}).keys())
            if months:
                return sorted(months)
        return []

    def _group_procurement(self, procurement: list[dict], key: str) -> list[dict]:
        if not procurement:
            return []
        order = {
            "紧急": 0,
            "关注": 1,
            "正常": 2,
            "充裕": 3,
            "立即采购": 0,
            "近期采购": 1,
            "按计划补货": 2,
            "人工复核": 3,
            "暂缓采购": 4,
        }
        groups: dict[str, dict] = {}
        for row in procurement:
            label = str(row.get(key, "") or "未分类")
            groups.setdefault(label, {key: label, "material_count": 0, "suggested_order_qty": 0, "forecast_total": 0, "sort_order": order.get(label, 99)})
            groups[label]["material_count"] += 1
            groups[label]["suggested_order_qty"] += int(row.get("suggested_order_qty", 0) or 0)
            groups[label]["forecast_total"] += float(row.get("forecast_total", 0) or 0)
        rows = list(groups.values())
        for row in rows:
            row["forecast_total"] = round(row["forecast_total"], 1)
        return sorted(rows, key=lambda x: x["sort_order"])

    def generate_html(self, output_path: str | Path, excel_path: str | Path | None = None) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(self._render_html(excel_path), encoding="utf-8")
        return output_path

    def generate_excel(self, output_path: str | Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            self._sheet_monthly().to_excel(writer, sheet_name="月度消耗趋势", index=False)
            self._sheet_doc_type().to_excel(writer, sheet_name="单据类型构成", index=False)
            self._sheet_materials().to_excel(writer, sheet_name="重点物料预测_全量", index=False)
            self._sheet_procurement().to_excel(writer, sheet_name="采购建议_全量", index=False)
            self._sheet_trend_summary().to_excel(writer, sheet_name="趋势统计", index=False)
            self._sheet_inactive_materials().to_excel(writer, sheet_name="近一年无发出物料", index=False)
            self._sheet_procurement_summary("priority", "优先级").to_excel(writer, sheet_name="采购优先级统计", index=False)
            self._sheet_procurement_summary("purchase_timing", "采购建议").to_excel(writer, sheet_name="采购标签统计", index=False)
            self._sheet_method_notes().to_excel(writer, sheet_name="口径说明", index=False)
            self._format_workbook(writer)
        return output_path

    def generate_summary_markdown(self) -> str:
        d = self.data
        summary = d.get("summary", {})
        procurement = summary.get("procurement_summary", {})
        top_items = d.get("material_forecasts", [])[:5]
        top_text = "\n".join(
            f"- {i.get('name', '')} ({i.get('code', '')}): 期间发出 {i.get('total_12m', 0):,.0f}, "
            f"未来3个月预测 {i.get('forecast_total', 0):,.0f}, 趋势 {i.get('trend', '')}"
            for i in top_items
        )
        return f"""### 库存分析摘要（{d.get("period", "")}）

组织：{d.get("organization", "")}

期间发出数量：{summary.get("total_qty_12m", 0):,.0f}
期间发出金额：{summary.get("total_out_amt_12m", 0):,.2f}
月均发出数量：{summary.get("avg_monthly", 0):,.0f}
未来3个月预测发出：{summary.get("forecast_total", 0):,.0f}

采购建议：紧急 {procurement.get("urgent_count", 0)} 个，关注 {procurement.get("watch_count", 0)} 个，建议采购总量 {procurement.get("total_suggested_order_qty", 0):,.0f}

重点物料：
{top_text}
"""

    def _sheet_monthly(self) -> pd.DataFrame:
        rows = []
        for month, row in sorted((self.data.get("monthly_trend") or {}).items()):
            rows.append(
                {
                    "月份": month,
                    "收入数量": row.get("income_qty", 0),
                    "发出数量": row.get("qty", 0),
                    "发出金额": row.get("amount", 0),
                }
            )
        return pd.DataFrame(rows)

    def _sheet_doc_type(self) -> pd.DataFrame:
        mapping = {
            "doc_type": "单据类型",
            "biz_type": "业务类型",
            "total_12m": "期间发出",
            "pct": "占比%",
            "avg_monthly": "月均",
            "ma3": "近3月均",
            "current_month_val": "最近月",
            "current_month_pct": "最近月占比%",
        }
        return pd.DataFrame(self.data.get("doc_type_breakdown") or []).rename(columns=mapping)

    def _sheet_materials(self) -> pd.DataFrame:
        months = self._forecast_months()
        rows = []
        for item in self.data.get("material_forecasts") or []:
            row = {
                "排名": item.get("rank"),
                "物料编码": item.get("code"),
                "物料名称": item.get("name"),
                "分组": item.get("group"),
                "期间发出": item.get("total_12m"),
                "月均": item.get("avg_monthly"),
                "近3月均": item.get("ma3"),
                "近6月消耗": item.get("recent_6m_total"),
                "未来合计": item.get("forecast_total"),
                "趋势": item.get("trend"),
                "预测方法": item.get("forecast_method"),
                "期末库存": item.get("end_qty"),
                "期末金额": item.get("end_amt"),
            }
            for month in months:
                row[month] = (item.get("forecast") or {}).get(month, 0)
            components = item.get("forecast_components") or {}
            row["预测下限"] = components.get("recent_floor", 0)
            rows.append(row)
        fixed = ["排名", "物料编码", "物料名称", "分组", "期间发出", "月均", "近3月均", "近6月消耗", *months, "未来合计", "趋势", "预测下限", "期末库存", "期末金额", "预测方法"]
        return pd.DataFrame(rows).reindex(columns=fixed)

    def _sheet_procurement(self) -> pd.DataFrame:
        mapping = {
            "rank": "排名",
            "code": "物料编码",
            "name": "物料名称",
            "group": "分组",
            "current_stock": "当前库存",
            "current_end_amt": "当前库存金额",
            "avg_monthly_consumption": "月均消耗",
            "avg_daily_consumption": "日均消耗",
            "days_of_supply": "可撑天数",
            "safety_stock": "安全库存",
            "reorder_point": "再订点",
            "forecast_total": "未来3月耗用",
            "suggested_order_qty": "建议采购量",
            "purchase_timing": "采购建议",
            "priority": "优先级",
        }
        return pd.DataFrame(self.data.get("procurement_suggestions") or []).rename(columns=mapping)

    def _sheet_inactive_materials(self) -> pd.DataFrame:
        mapping = {
            "code": "物料编码",
            "name": "物料名称",
            "group": "分组",
            "end_qty": "期末库存",
            "end_amt": "期末金额",
            "income_total_12m": "期间收入数量",
            "income_amt_total_12m": "期间收入金额",
        }
        return pd.DataFrame(self.data.get("inactive_materials") or []).rename(columns=mapping)

    def _sheet_trend_summary(self) -> pd.DataFrame:
        return pd.DataFrame(self.data.get("trend_summary") or []).rename(
            columns={"trend": "趋势", "material_count": "物料数量", "total_12m": "期间发出", "forecast_total": "未来3月预测"}
        )

    def _sheet_procurement_summary(self, key: str, label: str) -> pd.DataFrame:
        source = "procurement_summary_by_priority" if key == "priority" else "procurement_summary_by_action"
        return pd.DataFrame(self.data.get(source) or []).drop(columns=["sort_order"], errors="ignore").rename(
            columns={key: label, "material_count": "物料数量", "suggested_order_qty": "建议采购量", "forecast_total": "预测消耗量"}
        ).drop(columns=["预测消耗量"], errors="ignore")

    def _sheet_method_notes(self) -> pd.DataFrame:
        rows = []
        notes = self.data.get("method_notes") or {}
        for category, lines in notes.items():
            for idx, line in enumerate(lines, 1):
                rows.append({"类别": "预测方法" if category == "forecast" else "采购建议", "序号": idx, "说明": line})
        return pd.DataFrame(rows)

    def _format_workbook(self, writer) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="17324D")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col_idx, column_cells in enumerate(ws.columns, 1):
                values = [str(c.value) for c in column_cells if c.value is not None]
                width = min(max([len(v) for v in values] + [8]) + 2, 42)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.auto_filter.ref = ws.dimensions

    def _render_html(self, excel_path: str | Path | None) -> str:
        data = dict(self.data)
        data["detail_excel"] = Path(excel_path).name if excel_path else ""
        data_json = json.dumps(data, ensure_ascii=False)
        title = f"库存分析报告 - {self.data.get('organization', '')}"
        return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    :root {{
      --ink:#202124; --muted:#5f6368; --line:#d8dee8; --bg:#f4f6f9; --card:#fff;
      --head:#17324d; --blue:#1f5f99; --cyan:#0b7f9f; --green:#207245; --amber:#9a5b00; --red:#b3261e;
    }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; background:var(--bg); color:var(--ink); font-family:"Microsoft YaHei","PingFang SC","Segoe UI",Arial,sans-serif; }}
    .page {{ max-width:1480px; margin:0 auto; padding:22px; }}
    .toolbar {{ display:flex; justify-content:flex-end; margin-bottom:12px; }}
    button {{ border:0; border-radius:6px; background:var(--head); color:#fff; padding:9px 14px; font-size:14px; cursor:pointer; }}
    header {{ background:#17324d; color:#fff; border-radius:8px; padding:24px 28px; margin-bottom:16px; }}
    h1 {{ font-size:26px; margin:0 0 8px; letter-spacing:0; }}
    h2 {{ font-size:18px; margin:0 0 14px; color:var(--head); }}
    h3 {{ font-size:15px; margin:18px 0 10px; color:var(--head); }}
    .meta {{ display:flex; flex-wrap:wrap; gap:20px; font-size:13px; opacity:.92; }}
    .grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(190px,1fr)); gap:12px; margin:16px 0; }}
    .kpi,.section {{ background:var(--card); border:1px solid var(--line); border-radius:8px; }}
    .kpi {{ padding:15px; border-left:4px solid var(--cyan); }}
    .kpi .label {{ color:var(--muted); font-size:12px; }}
    .kpi .value {{ font-size:24px; font-weight:700; margin-top:5px; }}
    .section {{ padding:20px; margin-bottom:16px; overflow:hidden; }}
    .table-wrap {{ overflow:auto; border:1px solid var(--line); border-radius:6px; }}
    table {{ border-collapse:collapse; width:100%; min-width:900px; font-size:12px; }}
    th,td {{ border-bottom:1px solid var(--line); padding:8px 9px; text-align:center; vertical-align:middle; white-space:normal; }}
    th {{ background:#edf2f7; color:#263746; font-weight:650; position:sticky; top:0; }}
    td.name {{ max-width:280px; }}
    .badge {{ display:inline-block; min-width:44px; border-radius:999px; padding:2px 9px; font-weight:650; font-size:12px; }}
    .badge-紧急 {{ color:var(--red); background:#fce8e6; }}
    .badge-关注 {{ color:var(--amber); background:#fff3d6; }}
    .badge-正常 {{ color:var(--blue); background:#e8f0fe; }}
    .badge-充裕 {{ color:var(--green); background:#e6f4ea; }}
    .bars {{ display:grid; gap:8px; margin:8px 0 14px; }}
    .bar {{ display:grid; grid-template-columns:120px 1fr 90px; align-items:center; gap:10px; font-size:12px; text-align:center; }}
    .track {{ background:#e4eaf1; height:20px; border-radius:4px; overflow:hidden; }}
    .fill {{ height:100%; background:linear-gradient(90deg,#0b7f9f,#1f5f99); }}
    .combo-chart {{ width:100%; max-width:1180px; height:auto; display:block; margin:6px auto 4px; }}
    .chart-caption {{ color:var(--muted); font-size:12px; text-align:center; margin-top:8px; }}
    .summary {{ line-height:1.9; color:#303134; }}
    .two-col {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; }}
    footer {{ color:var(--muted); text-align:center; font-size:12px; margin:22px 0 4px; }}
    @media (max-width:900px) {{ .two-col {{ grid-template-columns:1fr; }} }}
    @media print {{
      @page {{ size:A4 landscape; margin:10mm; }}
      body {{ background:white; }}
      .page {{ padding:0; max-width:none; }}
      .toolbar {{ display:none; }}
      header {{ padding:14px 18px; margin-bottom:8px; break-inside:avoid; print-color-adjust:exact; -webkit-print-color-adjust:exact; }}
      h1 {{ font-size:20px; margin-bottom:5px; }}
      h2 {{ font-size:15px; margin-bottom:8px; break-after:avoid; }}
      h3 {{ font-size:13px; margin:10px 0 6px; break-after:avoid; }}
      .meta {{ gap:12px; font-size:11px; }}
      .grid {{ grid-template-columns:repeat(6,minmax(0,1fr)); gap:6px; margin:8px 0; }}
      .kpi {{ padding:8px 9px; border-left-width:3px; break-inside:avoid; print-color-adjust:exact; -webkit-print-color-adjust:exact; }}
      .kpi .label {{ font-size:10px; }}
      .kpi .value {{ font-size:15px; margin-top:3px; line-height:1.2; }}
      .section {{ padding:12px; margin-bottom:8px; overflow:visible; break-inside:auto; page-break-inside:auto; print-color-adjust:exact; -webkit-print-color-adjust:exact; }}
      .combo-chart {{ max-height:270px; }}
      table {{ font-size:10px; }}
      th,td {{ padding:5px 6px; }}
      tr {{ break-inside:avoid; page-break-inside:avoid; }}
      footer {{ margin-top:10px; font-size:10px; }}
    }}
  </style>
</head>
<body>
  <div class="page">
    <div class="toolbar"><button onclick="window.print()">打印 / 另存为 PDF</button></div>
    <div id="app"></div>
  </div>
  <script>const analysisData = {data_json};</script>
  <script>{self._client_script()}</script>
</body>
</html>"""

    def _client_script(self) -> str:
        return r"""
const d = analysisData;
const app = document.getElementById('app');
const fmt = (n, digits=0) => Number(n || 0).toLocaleString('zh-CN', {maximumFractionDigits: digits, minimumFractionDigits: digits});
const el = (tag, attrs={}, ...kids) => {
  const node = document.createElement(tag);
  Object.entries(attrs || {}).forEach(([k,v]) => k === 'class' ? node.className = v : node.setAttribute(k, v));
  kids.flat().forEach(k => node.append(k instanceof Node ? k : document.createTextNode(k ?? '')));
  return node;
};
const row = (cells, cls='') => el('tr', {class:cls}, cells.map((c,i) => el('td', {class:i===2?'name':''}, c)));
const table = (heads, rows) => el('div', {class:'table-wrap'}, el('table', {}, el('thead', {}, el('tr', {}, heads.map(h => el('th', {}, h)))), el('tbody', {}, rows)));
const s = d.summary || {};
const p = s.procurement_summary || {};
const forecastMonths = resolveForecastMonths();

app.append(el('header', {},
  el('h1', {}, '库存收发存分析报告'),
  el('div', {class:'meta'}, `组织：${d.organization || ''}`, `期间：${d.period || ''}`, `生成时间：${new Date().toLocaleString('zh-CN')}`)
));

app.append(el('div', {class:'grid'},
  kpi('期间收入数量', fmt(s.total_income_12m), '件'),
  kpi('期间发出数量', fmt(s.total_qty_12m), '件'),
  kpi('月均发出数量', fmt(s.avg_monthly), '件'),
  kpi('未来3个月预测消耗', fmt(s.material_forecast_total), '件'),
  kpi('紧急采购物料', fmt(p.urgent_count), '个'),
  kpi('建议采购总量', fmt(p.total_suggested_order_qty), '件')
));

monthlySection();
docTypeSection();
materialSection();
forecastSection();
procurementSection();
methodSection();
footer();

function resolveForecastMonths() {
  if (Array.isArray(d.forecast_months) && d.forecast_months.length) return d.forecast_months;
  const total = d.total_forecast || {};
  for (const key of ['ensemble', 'holt', 'recent_ma', 'long_avg']) {
    const months = Object.keys(total[key] || {});
    if (months.length) return months.sort();
  }
  const firstMaterial = (d.material_forecasts || []).find(x => x.forecast && Object.keys(x.forecast).length);
  return firstMaterial ? Object.keys(firstMaterial.forecast).sort() : [];
}

function kpi(label, value, unit) {
  return el('div', {class:'kpi'}, el('div', {class:'label'}, label), el('div', {class:'value'}, `${value} ${unit}`));
}

function bars(data, labelKey, valueKey, color='#1f5f99') {
  const max = Math.max(...data.map(x => x[valueKey] || 0), 1);
  return el('div', {class:'bars'}, data.map(x => {
    const val = x[valueKey] || 0;
    return el('div', {class:'bar'}, el('div', {}, x[labelKey]), el('div', {class:'track'}, el('div', {class:'fill', style:`width:${Math.max(val/max*100,1)}%;background:${color}`})), el('div', {}, fmt(val)));
  }));
}

function monthlySection() {
  const months = Object.keys(d.monthly_trend || {}).sort();
  const data = months.map(m => ({month:m, qty:d.monthly_trend[m].qty || 0, amount:d.monthly_trend[m].amount || 0}));
  app.append(section('一、月度消耗趋势', comboChart(data), el('div', {class:'chart-caption'}, '柱形为发出数量，折线为发出金额；左右双轴分别对应数量与金额。')));
}

function docTypeSection() {
  const rows = (d.doc_type_breakdown || []).map(x => row([x.doc_type, x.biz_type, fmt(x.total_12m), `${fmt(x.pct,1)}%`, fmt(x.avg_monthly), fmt(x.ma3), fmt(x.current_month_val), `${fmt(x.current_month_pct,1)}%`]));
  app.append(section('二、单据类型消耗构成', table(['单据类型','业务类型','期间发出','占比','月均','近3月均','最近月','最近月占比'], rows)));
}

function materialSection() {
  const inactive = d.inactive_material_summary || {};
  const rows = (d.material_forecasts || []).slice(0, 20).map(x => row([x.rank, x.code, x.name, x.group, fmt(x.total_12m), fmt(x.avg_monthly), fmt(x.ma3), ...forecastMonths.map(m => fmt((x.forecast || {})[m])), fmt(x.forecast_total), x.trend]));
  app.append(section('三、重点物料消耗预测',
    el('h3', {}, '物料消耗趋势统计图'),
    bars(d.trend_summary || [], 'trend', 'material_count', '#0b7f9f'),
    el('h3', {}, '近一年无发出物料统计'),
    inactiveStats(inactive),
    el('h3', {}, '期间发出量TOP20消耗预测'),
    table(['排名','物料编码','物料名称','分组','期间发出','月均','近3月均', ...forecastMonths, '未来合计','趋势'], rows)
  ));
}

function forecastSection() {
  const rows = (d.trend_summary || []).map(x => row([x.trend, fmt(x.material_count), fmt(x.total_12m), fmt(x.forecast_total)]));
  app.append(section('四、预测消耗汇总', table(['趋势','物料数量','期间发出','未来3月预测消耗'], rows)));
}

function procurementSection() {
  const rows = (d.procurement_suggestions || []).slice(0, 20).map(x => {
    const badge = el('span', {class:`badge badge-${x.priority}`}, x.priority);
    return row([x.rank, x.code, x.name, fmt(x.current_stock), fmt(x.avg_monthly_consumption,1), fmt(x.days_of_supply,1), fmt(x.safety_stock,1), fmt(x.reorder_point,1), fmt(x.forecast_total,1), fmt(x.suggested_order_qty), x.purchase_timing, badge]);
  });
  const priorityRows = (d.procurement_summary_by_priority || []).map(x => row([x.priority, fmt(x.material_count), fmt(x.suggested_order_qty)]));
  app.append(section('五、采购建议',
    el('h3', {}, '按优先级统计'),
    table(['优先级','物料数量','建议采购量'], priorityRows),
    el('h3', {}, '建议采购量TOP20'),
    table(['排名','物料编码','物料名称','当前库存','月均消耗','可撑天数','安全库存','再订点','未来3月耗用','建议采购量','采购建议','优先级'], rows)
  ));
}

function methodSection() {
  const notes = d.method_notes || {};
  app.append(section('六、分析摘要与口径说明', el('div', {class:'summary'},
    el('h3', {}, '分析摘要'),
    `期间发出数量 ${fmt(s.total_qty_12m)} 件，月均 ${fmt(s.avg_monthly)} 件；峰值月 ${s.peak_month || ''}，低谷月 ${s.trough_month || ''}。`,
    el('br'), `未来3个月预测消耗合计 ${fmt(s.material_forecast_total)} 件，建议采购总量 ${fmt(p.total_suggested_order_qty)} 件。预测消耗用于估计未来需求，建议采购量会进一步扣减当前库存并补足安全库存。`,
    el('br'), `趋势判断为 ${s.trend_direction || ''}。`,
    el('br'), `采购建议：紧急 ${fmt(p.urgent_count)} 个，关注 ${fmt(p.watch_count)} 个，建议采购总量 ${fmt(p.total_suggested_order_qty)} 件。`,
    el('br'), d.rebuilt_from_legacy_json ? '提示：本报告由历史 analysis_result.json 重建，单物料预测值沿用历史结果；如需应用最新预测算法，请从金蝶导出或已有 Excel 重新运行分析。' : '本报告基于源数据重新计算，已应用当前预测与采购建议口径。',
    el('h3', {}, '预测方法'),
    el('ol', {}, (notes.forecast || []).map(x => el('li', {}, x))),
    el('h3', {}, '采购建议口径'),
    el('ol', {}, (notes.procurement || []).map(x => el('li', {}, x)))
  )));
}

function section(title, ...children) {
  return el('section', {class:'section'}, el('h2', {}, title), ...children);
}
function comboChart(data) {
  const width = 1120, height = 360;
  const margin = {left:72, right:86, top:28, bottom:58};
  const plotW = width - margin.left - margin.right;
  const plotH = height - margin.top - margin.bottom;
  const maxQty = Math.max(...data.map(x => x.qty), 1);
  const maxAmt = Math.max(...data.map(x => x.amount), 1);
  const step = plotW / Math.max(data.length, 1);
  const barW = Math.max(12, step * 0.48);
  const x = i => margin.left + step * i + step / 2;
  const yQty = v => margin.top + plotH - (v / maxQty) * plotH;
  const yAmt = v => margin.top + plotH - (v / maxAmt) * plotH;
  const svg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
  svg.setAttribute('viewBox', `0 0 ${width} ${height}`);
  svg.setAttribute('class', 'combo-chart');
  const add = (tag, attrs, text) => {
    const n = document.createElementNS('http://www.w3.org/2000/svg', tag);
    Object.entries(attrs || {}).forEach(([k,v]) => n.setAttribute(k, v));
    if (text != null) n.textContent = text;
    svg.appendChild(n);
    return n;
  };
  for (let i=0; i<=4; i++) {
    const y = margin.top + plotH * i / 4;
    add('line', {x1:margin.left, y1:y, x2:width-margin.right, y2:y, stroke:'#e4eaf1'});
    add('text', {x:margin.left-10, y:y+4, 'text-anchor':'end', 'font-size':'11', fill:'#5f6368'}, fmt(maxQty*(1-i/4)));
    add('text', {x:width-margin.right+10, y:y+4, 'text-anchor':'start', 'font-size':'11', fill:'#5f6368'}, fmt(maxAmt*(1-i/4)/10000,1)+'万');
  }
  data.forEach((d,i) => {
    const cx = x(i);
    const barH = margin.top + plotH - yQty(d.qty);
    add('rect', {x:cx-barW/2, y:yQty(d.qty), width:barW, height:barH, fill:'#1f5f99', rx:3});
    add('text', {x:cx, y:Math.min(margin.top + plotH - 5, yQty(d.qty) + 14), 'text-anchor':'middle', 'font-size':'10', fill:'#ffffff'}, fmt(d.qty));
    add('text', {x:cx, y:height-32, 'text-anchor':'middle', 'font-size':'10', fill:'#5f6368'}, d.month.replace('-', '.'));
  });
  const points = data.map((d,i) => `${x(i)},${yAmt(d.amount)}`).join(' ');
  add('polyline', {points, fill:'none', stroke:'#b06000', 'stroke-width':3});
  data.forEach((d,i) => {
    add('circle', {cx:x(i), cy:yAmt(d.amount), r:4, fill:'#b06000'});
    add('text', {x:x(i), y:yAmt(d.amount)-8, 'text-anchor':'middle', 'font-size':'10', fill:'#9a5b00'}, fmt(d.amount/10000,1)+'万');
  });
  add('line', {x1:margin.left, y1:margin.top+plotH, x2:width-margin.right, y2:margin.top+plotH, stroke:'#8aa0b5'});
  add('text', {x:margin.left, y:18, 'font-size':'12', fill:'#1f5f99'}, '发出数量');
  add('text', {x:width-margin.right, y:18, 'text-anchor':'end', 'font-size':'12', fill:'#b06000'}, '发出金额');
  add('text', {x:width/2, y:height-8, 'text-anchor':'middle', 'font-size':'11', fill:'#5f6368'}, '月份');
  return svg;
}
function inactiveStats(inactive) {
  if (d.rebuilt_from_legacy_json && !(d.inactive_materials || []).length) {
    return el('div', {class:'summary'}, '当前报告由历史 analysis_result.json 重建，旧结果未保存近一年无发出物料清单；请从源 Excel 或金蝶数据重新运行分析后查看该统计。');
  }
  if (!inactive.material_count) {
    return el('div', {class:'summary'}, '近一年无发出物料为 0 个。');
  }
  return table(['指标','数值'], [
    row(['无发出物料数量', fmt(inactive.material_count)]),
    row(['其中仍有期末库存', fmt(inactive.stocked_count)]),
    row(['期末库存数量', fmt(inactive.end_qty)]),
    row(['期末库存金额', fmt(inactive.end_amt,2)])
  ]);
}
function sectionLite(title, ...children) {
  return el('div', {}, el('h3', {}, title), ...children);
}
function footer() {
  const detail = d.detail_excel ? `全量明细参考${d.detail_excel}；` : '';
  app.append(el('footer', {}, `数据来源：金蝶云星空；${detail}报告生成：KingdeeDataAnalyzer；${new Date().toLocaleString('zh-CN')}`));
}
"""

    def _forecast_months(self) -> list[str]:
        return list(self.data.get("forecast_months") or self._infer_forecast_months(self.data))
