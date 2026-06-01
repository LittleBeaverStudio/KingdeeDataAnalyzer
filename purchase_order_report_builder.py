from __future__ import annotations

import html
import json
from pathlib import Path

import pandas as pd


class PurchaseOrderReportBuilder:
    """Generate purchase order execution HTML report and Excel detail workbook."""

    def __init__(self, analysis_result: dict):
        self.data = dict(analysis_result)

    def generate_html(self, output_path: str | Path, excel_path: str | Path | None = None) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(self._render_html(excel_path), encoding="utf-8")
        return output_path

    def generate_excel(self, output_path: str | Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            self._summary_sheet().to_excel(writer, sheet_name="分析摘要", index=False)
            self._monthly_sheet().to_excel(writer, sheet_name="月度执行趋势", index=False)
            self._group_sheet("supplier_unsettled", "供应商").to_excel(writer, sheet_name="供应商未结算排行", index=False)
            self._group_sheet("material_execution", "物料").to_excel(writer, sheet_name="物料执行排行", index=False)
            self._group_sheet("purchaser_execution", "采购组织").to_excel(writer, sheet_name="采购组织执行", index=False)
            self._detail_sheet("overdue_unreceived").to_excel(writer, sheet_name="逾期未收料明细", index=False)
            self._detail_sheet("detail_rows").to_excel(writer, sheet_name="采购订单明细_全量", index=False)
            self._notes_sheet().to_excel(writer, sheet_name="口径说明", index=False)
            self._format_workbook(writer)
        return output_path

    def generate_summary_markdown(self) -> str:
        s = self.data.get("summary", {})
        return f"""### 采购订单分析摘要（{self.data.get("period", "")}）

组织：{self.data.get("organization", "")}

订单数：{s.get("order_count", 0):,}
供应商数：{s.get("supplier_count", 0):,}
价税合计：{s.get("order_amount", 0):,.2f}
应付金额：{s.get("payable_amount", 0):,.2f}
已结算金额：{s.get("settled_amount", 0):,.2f}
未结算金额：{s.get("unsettled_amount", 0):,.2f}
"""

    def _summary_sheet(self) -> pd.DataFrame:
        s = self.data.get("summary", {})
        rows = [
            ("订单数", s.get("order_count", 0)),
            ("订单行数", s.get("line_count", 0)),
            ("供应商数", s.get("supplier_count", 0)),
            ("物料数", s.get("material_count", 0)),
            ("订货数量", s.get("order_qty", 0)),
            ("收料数量", s.get("receive_qty", 0)),
            ("入库数量", s.get("stockin_qty", 0)),
            ("价税合计", s.get("order_amount", 0)),
            ("应付金额", s.get("payable_amount", 0)),
            ("已结算金额", s.get("settled_amount", 0)),
            ("未结算金额", s.get("unsettled_amount", 0)),
            ("未结算占应付比", f"{s.get('unsettled_ratio', 0)}%"),
            ("收料率", f"{s.get('receive_rate', 0)}%"),
            ("入库率", f"{s.get('stockin_rate', 0)}%"),
            ("结算率", f"{s.get('settlement_rate', 0)}%"),
            ("逾期未收料行数", s.get("overdue_line_count", 0)),
            ("逾期未收料数量", s.get("overdue_unreceived_qty", 0)),
        ]
        return pd.DataFrame(rows, columns=["指标", "数值"])

    def _monthly_sheet(self) -> pd.DataFrame:
        return pd.DataFrame(self.data.get("monthly_trend") or []).rename(
            columns={
                "month": "月份",
                "order_amount": "价税合计",
                "payable_amount": "应付金额",
                "settled_amount": "已结算金额",
                "unsettled_amount": "未结算金额",
                "order_qty": "订货数量",
                "receive_qty": "收料数量",
            }
        )

    def _group_sheet(self, key: str, label: str) -> pd.DataFrame:
        return pd.DataFrame(self.data.get(key) or []).rename(
            columns={
                "name": label,
                "order_count": "订单数",
                "line_count": "行数",
                "order_qty": "订货数量",
                "receive_qty": "收料数量",
                "stockin_qty": "入库数量",
                "order_amount": "价税合计",
                "payable_amount": "应付金额",
                "settled_amount": "已结算金额",
                "unsettled_amount": "未结算金额",
                "unsettled_ratio": "未结算占比%",
                "receive_rate": "收料率%",
                "unreceived_qty": "未收料数量",
            }
        )

    def _detail_sheet(self, key: str) -> pd.DataFrame:
        return pd.DataFrame(self.data.get(key) or []).rename(columns=self._detail_mapping())

    def _notes_sheet(self) -> pd.DataFrame:
        return pd.DataFrame([{"序号": idx, "说明": text} for idx, text in enumerate(self.data.get("method_notes") or [], 1)])

    def _detail_mapping(self) -> dict:
        return {
            "purchase_org": "采购组织",
            "order_no": "订单编号",
            "order_date": "订单日期",
            "supplier": "供应商",
            "material": "物料",
            "delivery_date": "交货日期",
            "currency": "币别",
            "order_qty": "订货数量",
            "receive_qty": "收料数量",
            "stockin_qty": "入库数量",
            "unreceived_qty": "未收料数量",
            "order_amount": "价税合计",
            "payable_amount": "应付金额",
            "settled_amount": "已结算金额",
            "unsettled_amount": "未结算金额",
            "receive_rate": "收料率",
            "settlement_rate": "结算率",
        }

    def _format_workbook(self, writer) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="17324D")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col_idx, column_cells in enumerate(ws.columns, 1):
                values = [str(c.value) for c in column_cells if c.value is not None]
                width = min(max([len(v) for v in values] + [8]) + 2, 42)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.auto_filter.ref = ws.dimensions

    def _render_html(self, excel_path: str | Path | None) -> str:
        data = dict(self.data)
        data["detail_excel"] = Path(excel_path).name if excel_path else ""
        data_json = json.dumps(data, ensure_ascii=False)
        title = f"采购订单分析报告 - {self.data.get('organization', '')}"
        return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    :root {{ --ink:#202124; --muted:#5f6368; --line:#d8dee8; --bg:#f4f6f9; --card:#fff; --head:#17324d; --blue:#1f5f99; --cyan:#0b7f9f; --amber:#9a5b00; --red:#b3261e; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; background:var(--bg); color:var(--ink); font-family:"Microsoft YaHei","PingFang SC","Segoe UI",Arial,sans-serif; }}
    .page {{ max-width:1480px; margin:0 auto; padding:22px; }}
    .toolbar {{ display:flex; justify-content:flex-end; margin-bottom:12px; }}
    button {{ border:0; border-radius:6px; background:var(--head); color:#fff; padding:9px 14px; font-size:14px; cursor:pointer; }}
    header {{ background:#17324d; color:#fff; border-radius:8px; padding:24px 28px; margin-bottom:16px; }}
    h1 {{ font-size:26px; margin:0 0 8px; letter-spacing:0; }}
    h2 {{ font-size:18px; margin:0 0 14px; color:var(--head); }}
    h3 {{ font-size:15px; margin:18px 0 10px; color:var(--head); }}
    .meta {{ display:flex; flex-wrap:wrap; gap:10px; font-size:13px; opacity:.92; }}
    .grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(190px,1fr)); gap:12px; margin:16px 0; }}
    .kpi,.section {{ background:var(--card); border:1px solid var(--line); border-radius:8px; }}
    .kpi {{ padding:15px; border-left:4px solid var(--cyan); }}
    .kpi.warn {{ border-left-color:var(--amber); }}
    .kpi.risk {{ border-left-color:var(--red); }}
    .kpi .label {{ color:var(--muted); font-size:12px; }}
    .kpi .value {{ font-size:24px; font-weight:700; margin-top:5px; }}
    .section {{ padding:20px; margin-bottom:16px; overflow:hidden; }}
    .table-wrap {{ overflow:auto; border:1px solid var(--line); border-radius:6px; }}
    table {{ border-collapse:collapse; width:100%; min-width:900px; font-size:12px; }}
    th,td {{ border-bottom:1px solid var(--line); padding:8px 9px; text-align:center; vertical-align:middle; white-space:normal; }}
    th {{ background:#edf2f7; color:#263746; font-weight:650; position:sticky; top:0; }}
    td.name {{ max-width:320px; }}
    .combo-chart {{ width:100%; max-width:1180px; height:auto; display:block; margin:6px auto 4px; }}
    .chart-caption {{ color:var(--muted); font-size:12px; text-align:center; margin-top:8px; }}
    .summary {{ line-height:1.9; color:#303134; }}
    footer {{ color:var(--muted); text-align:center; font-size:12px; margin:22px 0 4px; }}
    @media print {{
      @page {{ size:A4 landscape; margin:10mm; }}
      body {{ background:white; }}
      .page {{ padding:0; max-width:none; }}
      .toolbar {{ display:none; }}
      header {{ padding:14px 18px; margin-bottom:8px; break-inside:avoid; print-color-adjust:exact; -webkit-print-color-adjust:exact; }}
      h1 {{ font-size:20px; margin-bottom:5px; }}
      h2 {{ font-size:15px; margin-bottom:8px; break-after:avoid; }}
      h3 {{ font-size:13px; margin:10px 0 6px; break-after:avoid; }}
      .meta {{ gap:12px; font-size:11px; }}
      .grid {{ grid-template-columns:repeat(6,minmax(0,1fr)); gap:6px; margin:8px 0; }}
      .kpi {{ padding:8px 9px; border-left-width:3px; break-inside:avoid; print-color-adjust:exact; -webkit-print-color-adjust:exact; }}
      .kpi .label {{ font-size:10px; }}
      .kpi .value {{ font-size:15px; margin-top:3px; line-height:1.2; }}
      .section {{ padding:12px; margin-bottom:8px; overflow:visible; break-inside:auto; page-break-inside:auto; print-color-adjust:exact; -webkit-print-color-adjust:exact; }}
      .combo-chart {{ max-height:270px; }}
      table {{ font-size:10px; }}
      th,td {{ padding:5px 6px; }}
      tr {{ break-inside:avoid; page-break-inside:avoid; }}
      footer {{ margin-top:10px; font-size:10px; }}
    }}
  </style>
</head>
<body>
  <div class="page">
    <div class="toolbar"><button onclick="window.print()">打印 / 另存为 PDF</button></div>
    <div id="app"></div>
  </div>
  <script>const analysisData = {data_json};</script>
  <script>{self._client_script()}</script>
</body>
</html>"""

    def _client_script(self) -> str:
        return r"""
const d = analysisData;
const app = document.getElementById('app');
const fmt = (n, digits=0) => Number(n || 0).toLocaleString('zh-CN', {maximumFractionDigits: digits, minimumFractionDigits: digits});
const money = n => fmt(n, 2);
const pct = n => `${fmt(n, 1)}%`;
const el = (tag, attrs={}, ...kids) => {
  const node = document.createElement(tag);
  Object.entries(attrs || {}).forEach(([k,v]) => k === 'class' ? node.className = v : node.setAttribute(k, v));
  kids.flat().forEach(k => node.append(k instanceof Node ? k : document.createTextNode(k ?? '')));
  return node;
};
const row = cells => el('tr', {}, cells.map((c,i) => el('td', {class:i===0?'name':''}, c)));
const table = (heads, rows) => el('div', {class:'table-wrap'}, el('table', {}, el('thead', {}, el('tr', {}, heads.map(h => el('th', {}, h)))), el('tbody', {}, rows)));
const s = d.summary || {};

app.append(el('header', {},
  el('h1', {}, '采购订单分析报告'),
  el('div', {class:'meta'}, `组织：${d.organization || '自动识别组织'}；期间：${d.period || ''}；生成时间：${new Date().toLocaleString('zh-CN')}`)
));

app.append(el('div', {class:'grid'},
  kpi('订单数', fmt(s.order_count), '单'),
  kpi('价税合计', money(s.order_amount), '元'),
  kpi('应付金额', money(s.payable_amount), '元'),
  kpi('已结算金额', money(s.settled_amount), '元'),
  kpi('未结算金额', money(s.unsettled_amount), '元', 'risk'),
  kpi('未结算占应付比', pct(s.unsettled_ratio), '', 'warn')
));

monthlySection();
supplierSection();
materialSection();
overdueSection();
methodSection();
footer();

function kpi(label, value, unit, cls='') {
  return el('div', {class:`kpi ${cls}`}, el('div', {class:'label'}, label), el('div', {class:'value'}, `${value} ${unit}`.trim()));
}

function section(title, ...children) {
  return el('section', {class:'section'}, el('h2', {}, title), ...children);
}

function monthlySection() {
  const data = d.monthly_trend || [];
  app.append(section('一、月度采购执行趋势',
    comboChart(data),
    el('div', {class:'chart-caption'}, '每月一组重叠柱：底层为应付金额，前景为未结算金额，使用同一金额坐标轴。')
  ));
}

function supplierSection() {
  const rows = (d.supplier_unsettled || []).slice(0, 20).map(x => row([x.name, fmt(x.order_count), money(x.order_amount), money(x.payable_amount), money(x.settled_amount), money(x.unsettled_amount), pct(x.unsettled_ratio), pct(x.receive_rate), fmt(x.unreceived_qty)]));
  app.append(section('二、供应商未结算排行TOP20', table(['供应商','订单数','价税合计','应付金额','已结算金额','未结算金额','未结算占比','收料率','未收料数量'], rows)));
}

function materialSection() {
  const rows = (d.material_execution || []).slice(0, 20).map(x => row([x.name, fmt(x.line_count), fmt(x.order_qty), fmt(x.receive_qty), fmt(x.stockin_qty), money(x.payable_amount), money(x.unsettled_amount), pct(x.unsettled_ratio), fmt(x.unreceived_qty)]));
  app.append(section('三、物料采购执行TOP20', table(['物料','行数','订货数量','收料数量','入库数量','应付金额','未结算金额','未结算占比','未收料数量'], rows)));
}

function overdueSection() {
  const rows = (d.overdue_unreceived || []).slice(0, 20).map(x => row([x.order_no, x.order_date, x.supplier, x.material, x.delivery_date, fmt(x.order_qty), fmt(x.receive_qty), fmt(x.unreceived_qty), money(x.unsettled_amount)]));
  const content = rows.length ? table(['订单编号','订单日期','供应商','物料','交货日期','订货数量','收料数量','未收料数量','未结算金额'], rows) : el('div', {class:'summary'}, '本期间未识别到逾期未收料订单行。');
  app.append(section('四、逾期未收料风险', content));
}

function methodSection() {
  const notes = d.method_notes || [];
  app.append(section('五、分析摘要与口径说明',
    el('div', {class:'summary'},
      el('h3', {}, '分析摘要'),
      `本期采购订单 ${fmt(s.order_count)} 单、${fmt(s.line_count)} 行，供应商 ${fmt(s.supplier_count)} 个，物料 ${fmt(s.material_count)} 个。`,
      el('br'), `价税合计 ${money(s.order_amount)} 元，应付金额 ${money(s.payable_amount)} 元，已结算金额 ${money(s.settled_amount)} 元，未结算金额 ${money(s.unsettled_amount)} 元。`,
      el('br'), `收料率 ${pct(s.receive_rate)}，入库率 ${pct(s.stockin_rate)}，结算率 ${pct(s.settlement_rate)}。逾期未收料行数 ${fmt(s.overdue_line_count)} 行，逾期未收料数量 ${fmt(s.overdue_unreceived_qty)}。`,
      el('h3', {}, '核心口径'),
      el('ol', {}, notes.map(x => el('li', {}, x))),
      el('h3', {}, '使用建议'),
      '优先查看未结算金额较高的供应商和物料，再结合逾期未收料清单判断是结算推进问题、到货执行问题，还是订单关闭口径需要复核。'
    )
  ));
}

function comboChart(data) {
  const width = 1120, height = 360;
  const margin = {left:86, right:32, top:30, bottom:58};
  const plotW = width - margin.left - margin.right;
  const plotH = height - margin.top - margin.bottom;
  const values = data.flatMap(x => [x.payable_amount || 0, x.unsettled_amount || 0]);
  const maxValue = Math.max(...values.map(v => Math.abs(v)), 1);
  const step = plotW / Math.max(data.length, 1);
  const baseBarW = Math.max(18, step * 0.48);
  const overlayBarW = Math.max(10, step * 0.24);
  const x = i => margin.left + step * i + step / 2;
  const zeroY = margin.top + plotH;
  const yVal = v => zeroY - (v / maxValue) * plotH * 0.92;
  const svg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
  svg.setAttribute('viewBox', `0 0 ${width} ${height}`);
  svg.setAttribute('class', 'combo-chart');
  const add = (tag, attrs, text) => {
    const n = document.createElementNS('http://www.w3.org/2000/svg', tag);
    Object.entries(attrs || {}).forEach(([k,v]) => n.setAttribute(k, v));
    if (text != null) n.textContent = text;
    svg.appendChild(n);
    return n;
  };
  for (let i=0; i<=4; i++) {
    const y = margin.top + plotH * i / 4;
    add('line', {x1:margin.left, y1:y, x2:width-margin.right, y2:y, stroke:'#e4eaf1'});
    add('text', {x:margin.left-10, y:y+4, 'text-anchor':'end', 'font-size':'11', fill:'#5f6368'}, fmt(maxValue*(1-i/4)/10000,1)+'万');
  }
  data.forEach((d,i) => {
    const cx = x(i);
    drawBar(cx - baseBarW / 2, d.payable_amount || 0, baseBarW, '#8fb8dd', false);
    drawBar(cx - overlayBarW / 2, d.unsettled_amount || 0, overlayBarW, '#b06000', true);
    add('text', {x:cx, y:height-32, 'text-anchor':'middle', 'font-size':'10', fill:'#5f6368'}, d.month);
  });
  add('line', {x1:margin.left, y1:zeroY, x2:width-margin.right, y2:zeroY, stroke:'#8aa0b5'});
  add('rect', {x:margin.left, y:12, width:12, height:8, fill:'#8fb8dd'});
  add('text', {x:margin.left+18, y:20, 'font-size':'12', fill:'#263746'}, '应付金额');
  add('rect', {x:margin.left+92, y:12, width:12, height:8, fill:'#b06000'});
  add('text', {x:margin.left+110, y:20, 'font-size':'12', fill:'#263746'}, '未结算金额');
  function drawBar(x0, value, width0, color, showLabel) {
    const y = yVal(value);
    const top = Math.min(y, zeroY);
    const h = Math.max(Math.abs(zeroY - y), 1);
    add('rect', {x:x0, y:top, width:width0, height:h, fill:color, rx:3});
    if (showLabel && Math.abs(value) > 0) {
      add('text', {x:x0+width0/2, y:top-6, 'text-anchor':'middle', 'font-size':'9', fill:color}, fmt(value/10000,1)+'万');
    }
  }
  return svg;
}

function footer() {
  const detail = d.detail_excel ? `全量明细参考${d.detail_excel}；` : '';
  app.append(el('footer', {}, `数据来源：金蝶云星空；${detail}报告生成：KingdeeDataAnalyzer；${new Date().toLocaleString('zh-CN')}`));
}
"""
