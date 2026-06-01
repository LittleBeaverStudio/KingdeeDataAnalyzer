from __future__ import annotations

import html
import json
from pathlib import Path

import pandas as pd


class SalesOutstockReportBuilder:
    """Generate sales outstock invoice tracking HTML report and Excel detail workbook."""

    def __init__(self, analysis_result: dict):
        self.data = dict(analysis_result)

    def generate_html(self, output_path: str | Path, excel_path: str | Path | None = None) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(self._render_html(excel_path), encoding="utf-8")
        return output_path

    def generate_excel(self, output_path: str | Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            self._summary_sheet().to_excel(writer, sheet_name="分析摘要", index=False)
            self._monthly_sheet().to_excel(writer, sheet_name="月度销售结算趋势", index=False)
            self._group_sheet("bill_type_summary", "单据类型").to_excel(writer, sheet_name="单据类型统计", index=False)
            self._group_sheet("salesperson_summary", "销售员").to_excel(writer, sheet_name="销售员统计", index=False)
            self._group_sheet("material_summary", "物料").to_excel(writer, sheet_name="物料统计", index=False)
            self._group_sheet("customer_summary", "客户").to_excel(writer, sheet_name="客户统计", index=False)
            self._detail_sheet("unsettled_details").to_excel(writer, sheet_name="未结算明细", index=False)
            self._detail_sheet("detail_rows").to_excel(writer, sheet_name="销售出库明细_全量", index=False)
            self._notes_sheet().to_excel(writer, sheet_name="口径说明", index=False)
            self._format_workbook(writer)
        return output_path

    def generate_summary_markdown(self) -> str:
        s = self.data.get("summary", {})
        return f"""### 销售出库分析摘要（{self.data.get("period", "")}）

组织：{self.data.get("organization", "")}

单据数：{s.get("bill_count", 0):,}
客户数：{s.get("customer_count", 0):,}
出库金额：{s.get("amount", 0):,.2f}
应收金额：{s.get("receivable_amount", 0):,.2f}
收款结算金额：{s.get("receipt_settlement_amount", 0):,.2f}
未结算金额：{s.get("unsettled_amount", 0):,.2f}
"""

    def _summary_sheet(self) -> pd.DataFrame:
        s = self.data.get("summary", {})
        rows = [
            ("单据数", s.get("bill_count", 0)),
            ("明细行数", s.get("line_count", 0)),
            ("客户数", s.get("customer_count", 0)),
            ("销售员数", s.get("salesperson_count", 0)),
            ("物料数", s.get("material_count", 0)),
            ("出库数量", s.get("qty", 0)),
            ("出库金额", s.get("amount", 0)),
            ("应收数量", s.get("receivable_qty", 0)),
            ("应收金额", s.get("receivable_amount", 0)),
            ("开票金额", s.get("invoice_amount", 0)),
            ("收款结算金额", s.get("receipt_settlement_amount", 0)),
            ("未结算金额", s.get("unsettled_amount", 0)),
            ("未开票金额", s.get("uninvoiced_amount", 0)),
            ("开票率", f"{s.get('invoice_rate', 0)}%"),
            ("结算率", f"{s.get('settlement_rate', 0)}%"),
            ("未结算占比", f"{s.get('unsettled_ratio', 0)}%"),
            ("平均单价", s.get("avg_price", 0)),
        ]
        return pd.DataFrame(rows, columns=["指标", "数值"])

    def _monthly_sheet(self) -> pd.DataFrame:
        return pd.DataFrame(self.data.get("monthly_trend") or []).rename(columns=self._monthly_mapping())

    def _group_sheet(self, key: str, label: str) -> pd.DataFrame:
        return pd.DataFrame(self.data.get(key) or []).rename(
            columns={
                "name": label,
                "bill_count": "单据数",
                "line_count": "行数",
                "qty": "出库数量",
                "amount": "出库金额",
                "receivable_amount": "应收金额",
                "invoice_amount": "开票金额",
                "receipt_settlement_amount": "收款结算金额",
                "unsettled_amount": "未结算金额",
                "uninvoiced_amount": "未开票金额",
                "invoice_rate": "开票率%",
                "settlement_rate": "结算率%",
                "unsettled_ratio": "未结算占比%",
            }
        )

    def _detail_sheet(self, key: str) -> pd.DataFrame:
        return pd.DataFrame(self.data.get(key) or []).rename(columns=self._detail_mapping())

    def _notes_sheet(self) -> pd.DataFrame:
        return pd.DataFrame([{"序号": idx, "说明": text} for idx, text in enumerate(self.data.get("method_notes") or [], 1)])

    def _monthly_mapping(self) -> dict:
        return {
            "month": "月份",
            "qty": "出库数量",
            "amount": "出库金额",
            "receivable_amount": "应收金额",
            "invoice_amount": "开票金额",
            "receipt_settlement_amount": "收款结算金额",
            "unsettled_amount": "未结算金额",
        }

    def _detail_mapping(self) -> dict:
        return {
            "sale_org": "销售组织",
            "bill_no": "单据编号",
            "bill_type": "单据类型",
            "bill_date": "日期",
            "salesperson": "销售员",
            "customer": "客户名称",
            "material": "物料名称",
            "qty": "数量",
            "price": "单价",
            "amount": "金额",
            "is_free": "是否赠品",
            "receivable_qty": "应收数量",
            "receivable_amount": "应收金额",
            "adjustment_amount": "调整金额",
            "invoice_qty": "开票数量",
            "invoice_amount": "开票金额",
            "receipt_settlement_amount": "收款结算金额",
            "settlement_adjustment": "结算调整金额",
            "special_writeoff": "特殊冲销金额",
            "unsettled_amount": "未结算金额",
            "uninvoiced_amount": "未开票金额",
            "invoice_rate": "开票率",
            "settlement_rate": "结算率",
        }

    def _format_workbook(self, writer) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="17324D")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col_idx, column_cells in enumerate(ws.columns, 1):
                values = [str(c.value) for c in column_cells if c.value is not None]
                width = min(max([len(v) for v in values] + [8]) + 2, 42)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.auto_filter.ref = ws.dimensions

    def _render_html(self, excel_path: str | Path | None) -> str:
        data = dict(self.data)
        data["detail_excel"] = Path(excel_path).name if excel_path else ""
        data_json = json.dumps(data, ensure_ascii=False)
        title = f"销售出库分析报告 - {self.data.get('organization', '')}"
        return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    :root {{ --ink:#202124; --muted:#5f6368; --line:#d8dee8; --bg:#f4f6f9; --card:#fff; --head:#17324d; --blue:#1f5f99; --cyan:#0b7f9f; --amber:#9a5b00; --red:#b3261e; --green:#207245; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; background:var(--bg); color:var(--ink); font-family:"Microsoft YaHei","PingFang SC","Segoe UI",Arial,sans-serif; }}
    .page {{ max-width:1480px; margin:0 auto; padding:22px; }}
    .toolbar {{ display:flex; justify-content:flex-end; margin-bottom:12px; }}
    button {{ border:0; border-radius:6px; background:var(--head); color:#fff; padding:9px 14px; font-size:14px; cursor:pointer; }}
    header {{ background:#17324d; color:#fff; border-radius:8px; padding:24px 28px; margin-bottom:16px; }}
    h1 {{ font-size:26px; margin:0 0 8px; letter-spacing:0; }}
    h2 {{ font-size:18px; margin:0 0 14px; color:var(--head); }}
    h3 {{ font-size:15px; margin:18px 0 10px; color:var(--head); }}
    .meta {{ display:flex; flex-wrap:wrap; gap:20px; font-size:13px; opacity:.92; }}
    .grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(190px,1fr)); gap:12px; margin:16px 0; }}
    .kpi,.section {{ background:var(--card); border:1px solid var(--line); border-radius:8px; }}
    .kpi {{ padding:15px; border-left:4px solid var(--cyan); }}
    .kpi.risk {{ border-left-color:var(--red); }}
    .kpi.warn {{ border-left-color:var(--amber); }}
    .kpi .label {{ color:var(--muted); font-size:12px; }}
    .kpi .value {{ font-size:24px; font-weight:700; margin-top:5px; }}
    .section {{ padding:20px; margin-bottom:16px; overflow:hidden; }}
    .table-wrap {{ overflow:auto; border:1px solid var(--line); border-radius:6px; }}
    table {{ border-collapse:collapse; width:100%; min-width:900px; font-size:12px; }}
    th,td {{ border-bottom:1px solid var(--line); padding:8px 9px; text-align:center; vertical-align:middle; white-space:normal; }}
    th {{ background:#edf2f7; color:#263746; font-weight:650; position:sticky; top:0; }}
    td.name {{ max-width:320px; }}
    .combo-chart {{ width:100%; max-width:1180px; height:auto; display:block; margin:6px auto 4px; }}
    .chart-caption {{ color:var(--muted); font-size:12px; text-align:center; margin-top:8px; }}
    .summary {{ line-height:1.9; color:#303134; }}
    footer {{ color:var(--muted); text-align:center; font-size:12px; margin:22px 0 4px; }}
    @media print {{
      @page {{ size:A4 landscape; margin:10mm; }}
      body {{ background:white; }}
      .page {{ padding:0; max-width:none; }}
      .toolbar {{ display:none; }}
      header {{ padding:14px 18px; margin-bottom:8px; break-inside:avoid; print-color-adjust:exact; -webkit-print-color-adjust:exact; }}
      h1 {{ font-size:20px; margin-bottom:5px; }}
      h2 {{ font-size:15px; margin-bottom:8px; break-after:avoid; }}
      h3 {{ font-size:13px; margin:10px 0 6px; break-after:avoid; }}
      .meta {{ gap:12px; font-size:11px; }}
      .grid {{ grid-template-columns:repeat(6,minmax(0,1fr)); gap:6px; margin:8px 0; }}
      .kpi {{ padding:8px 9px; border-left-width:3px; break-inside:avoid; print-color-adjust:exact; -webkit-print-color-adjust:exact; }}
      .kpi .label {{ font-size:10px; }}
      .kpi .value {{ font-size:15px; margin-top:3px; line-height:1.2; }}
      .section {{ padding:12px; margin-bottom:8px; overflow:visible; break-inside:auto; page-break-inside:auto; print-color-adjust:exact; -webkit-print-color-adjust:exact; }}
      .combo-chart {{ max-height:270px; }}
      table {{ font-size:10px; }}
      th,td {{ padding:5px 6px; }}
      tr {{ break-inside:avoid; page-break-inside:avoid; }}
      footer {{ margin-top:10px; font-size:10px; }}
    }}
  </style>
</head>
<body>
  <div class="page">
    <div class="toolbar"><button onclick="window.print()">打印 / 另存为 PDF</button></div>
    <div id="app"></div>
  </div>
  <script>const analysisData = {data_json};</script>
  <script>{self._client_script()}</script>
</body>
</html>"""

    def _client_script(self) -> str:
        return r"""
const d = analysisData;
const app = document.getElementById('app');
const fmt = (n, digits=0) => Number(n || 0).toLocaleString('zh-CN', {maximumFractionDigits: digits, minimumFractionDigits: digits});
const money = n => fmt(n, 2);
const pct = n => `${fmt(n, 1)}%`;
const el = (tag, attrs={}, ...kids) => {
  const node = document.createElement(tag);
  Object.entries(attrs || {}).forEach(([k,v]) => k === 'class' ? node.className = v : node.setAttribute(k, v));
  kids.flat().forEach(k => node.append(k instanceof Node ? k : document.createTextNode(k ?? '')));
  return node;
};
const row = cells => el('tr', {}, cells.map((c,i) => el('td', {class:i===0?'name':''}, c)));
const table = (heads, rows) => el('div', {class:'table-wrap'}, el('table', {}, el('thead', {}, el('tr', {}, heads.map(h => el('th', {}, h)))), el('tbody', {}, rows)));
const s = d.summary || {};

app.append(el('header', {},
  el('h1', {}, '销售出库分析报告'),
  el('div', {class:'meta'}, `组织：${d.organization || ''}`, `期间：${d.period || ''}`, `生成时间：${new Date().toLocaleString('zh-CN')}`)
));

app.append(el('div', {class:'grid'},
  kpi('单据数', fmt(s.bill_count), '单'),
  kpi('出库金额', money(s.amount), '元'),
  kpi('应收金额', money(s.receivable_amount), '元'),
  kpi('开票金额', money(s.invoice_amount), '元'),
  kpi('收款结算金额', money(s.receipt_settlement_amount), '元'),
  kpi('未结算金额', money(s.unsettled_amount), '元', 'risk')
));

monthlySection();
billTypeSection();
salespersonSection();
materialSection();
unsettledSection();
methodSection();
footer();

function kpi(label, value, unit, cls='') {
  return el('div', {class:`kpi ${cls}`}, el('div', {class:'label'}, label), el('div', {class:'value'}, `${value} ${unit}`.trim()));
}
function section(title, ...children) {
  return el('section', {class:'section'}, el('h2', {}, title), ...children);
}
function monthlySection() {
  const data = d.monthly_trend || [];
  app.append(section('一、月度销售与结算趋势',
    comboChart(data),
    el('div', {class:'chart-caption'}, '柱形为应收金额，折线为未结算金额；右侧轴对应未结算金额。')
  ));
}
function billTypeSection() {
  const rows = (d.bill_type_summary || []).slice(0, 20).map(x => groupRow(x));
  app.append(section('二、单据类型统计', table(groupHeads('单据类型'), rows)));
}
function salespersonSection() {
  const rows = (d.salesperson_summary || []).slice(0, 20).map(x => groupRow(x));
  app.append(section('三、销售员统计TOP20', table(groupHeads('销售员'), rows)));
}
function materialSection() {
  const rows = (d.material_summary || []).slice(0, 20).map(x => groupRow(x));
  app.append(section('四、物料统计TOP20', table(groupHeads('物料'), rows)));
}
function unsettledSection() {
  const rows = (d.unsettled_details || []).slice(0, 20).map(x => row([x.bill_no, x.bill_date, x.bill_type, x.salesperson, x.customer, fmt(x.line_count), fmt(x.qty), money(x.receivable_amount), money(x.invoice_amount), money(x.receipt_settlement_amount), money(x.unsettled_amount), pct((x.settlement_rate || 0) * 100)]));
  const content = rows.length ? table(['单据编号','日期','单据类型','销售员','客户','行数','数量','应收金额','开票金额','收款结算金额','未结算金额','结算率'], rows) : el('div', {class:'summary'}, '本期间未识别到销售出库结算明细。');
  app.append(section('五、未结算金额明细TOP20', content));
}
function methodSection() {
  const notes = d.method_notes || [];
  app.append(section('六、分析摘要与口径说明',
    el('div', {class:'summary'},
      el('h3', {}, '分析摘要'),
      `本期销售出库单据 ${fmt(s.bill_count)} 单、${fmt(s.line_count)} 行，客户 ${fmt(s.customer_count)} 个，销售员 ${fmt(s.salesperson_count)} 个，物料 ${fmt(s.material_count)} 个。`,
      el('br'), `出库金额 ${money(s.amount)} 元，应收金额 ${money(s.receivable_amount)} 元，开票金额 ${money(s.invoice_amount)} 元，收款结算金额 ${money(s.receipt_settlement_amount)} 元，未结算金额 ${money(s.unsettled_amount)} 元。`,
      el('br'), `开票率 ${pct(s.invoice_rate)}，结算率 ${pct(s.settlement_rate)}，未结算占比 ${pct(s.unsettled_ratio)}。`,
      el('h3', {}, '核心口径'),
      el('ol', {}, notes.map(x => el('li', {}, x))),
      el('h3', {}, '使用建议'),
      '优先查看未结算金额绝对值较高的单据、销售员、物料和客户，再结合单据类型判断是业务未结算、退货影响，还是收款结算回写口径需要复核。'
    )
  ));
}
function groupHeads(first) {
  return [first,'单据数','行数','数量','出库金额','应收金额','开票金额','收款结算金额','未结算金额','未开票金额','开票率','结算率','未结算占比'];
}
function groupRow(x) {
  return row([x.name, fmt(x.bill_count), fmt(x.line_count), fmt(x.qty), money(x.amount), money(x.receivable_amount), money(x.invoice_amount), money(x.receipt_settlement_amount), money(x.unsettled_amount), money(x.uninvoiced_amount), pct(x.invoice_rate), pct(x.settlement_rate), pct(x.unsettled_ratio)]);
}
function comboChart(data) {
  const width = 1120, height = 360;
  const margin = {left:72, right:86, top:28, bottom:58};
  const plotW = width - margin.left - margin.right;
  const plotH = height - margin.top - margin.bottom;
  const maxRec = Math.max(...data.map(x => x.receivable_amount || 0), 1);
  const maxUnsettled = Math.max(...data.map(x => Math.abs(x.unsettled_amount || 0)), 1);
  const step = plotW / Math.max(data.length, 1);
  const barW = Math.max(12, step * 0.48);
  const x = i => margin.left + step * i + step / 2;
  const yRec = v => margin.top + plotH - (v / maxRec) * plotH;
  const yUn = v => margin.top + plotH / 2 - (v / maxUnsettled) * plotH * 0.45;
  const svg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
  svg.setAttribute('viewBox', `0 0 ${width} ${height}`);
  svg.setAttribute('class', 'combo-chart');
  const add = (tag, attrs, text) => {
    const n = document.createElementNS('http://www.w3.org/2000/svg', tag);
    Object.entries(attrs || {}).forEach(([k,v]) => n.setAttribute(k, v));
    if (text != null) n.textContent = text;
    svg.appendChild(n);
    return n;
  };
  for (let i=0; i<=4; i++) {
    const y = margin.top + plotH * i / 4;
    add('line', {x1:margin.left, y1:y, x2:width-margin.right, y2:y, stroke:'#e4eaf1'});
    add('text', {x:margin.left-10, y:y+4, 'text-anchor':'end', 'font-size':'11', fill:'#5f6368'}, fmt(maxRec*(1-i/4)/10000,1)+'万');
    const rightValue = maxUnsettled * (1 - i / 2);
    add('text', {x:width-margin.right+10, y:y+4, 'text-anchor':'start', 'font-size':'11', fill:'#5f6368'}, fmt(rightValue/10000,1)+'万');
  }
  data.forEach((d,i) => {
    const cx = x(i);
    add('rect', {x:cx-barW/2, y:yRec(d.receivable_amount), width:barW, height:margin.top+plotH-yRec(d.receivable_amount), fill:'#1f5f99', rx:3});
    add('text', {x:cx, y:height-32, 'text-anchor':'middle', 'font-size':'10', fill:'#5f6368'}, d.month);
  });
  const points = data.map((d,i) => `${x(i)},${yUn(d.unsettled_amount || 0)}`).join(' ');
  add('polyline', {points, fill:'none', stroke:'#b06000', 'stroke-width':3});
  data.forEach((d,i) => {
    add('circle', {cx:x(i), cy:yUn(d.unsettled_amount || 0), r:4, fill:'#b06000'});
    add('text', {x:x(i), y:yUn(d.unsettled_amount || 0)-8, 'text-anchor':'middle', 'font-size':'10', fill:'#9a5b00'}, fmt((d.unsettled_amount || 0)/10000,1)+'万');
  });
  add('line', {x1:margin.left, y1:margin.top+plotH, x2:width-margin.right, y2:margin.top+plotH, stroke:'#8aa0b5'});
  add('text', {x:margin.left, y:18, 'font-size':'12', fill:'#1f5f99'}, '应收金额');
  add('line', {x1:margin.left, y1:margin.top+plotH/2, x2:width-margin.right, y2:margin.top+plotH/2, stroke:'#c7d0da', 'stroke-dasharray':'4 4'});
  add('text', {x:width-margin.right, y:18, 'text-anchor':'end', 'font-size':'12', fill:'#b06000'}, '未结算金额');
  return svg;
}
function footer() {
  const detail = d.detail_excel ? `全量明细参考${d.detail_excel}；` : '';
  app.append(el('footer', {}, `数据来源：金蝶云星空；${detail}报告生成：KingdeeDataAnalyzer；${new Date().toLocaleString('zh-CN')}`));
}
"""
